"""
main.py — сервис контроля лимитов Битрикс24
Запуск: uvicorn main:app --host 0.0.0.0 --port 8001
"""
import os, logging
from datetime import datetime
from typing import Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, Depends, HTTPException, Request, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy.orm import Session
import openpyxl
from io import BytesIO

from logging_config import setup_logging
setup_logging()   # ← первым делом, до всех импортов

from database import (
    get_db, init_db, seed_test_data, SessionLocal,
    Department, DepartmentLimit, LimitTransaction, ElementTracking, Setting
)
from bitrix import bitrix

logger = logging.getLogger(__name__)
WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET", "change-me-secret")


# ────────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("=" * 60)
    logger.info("[STARTUP] Сервис контроля лимитов стартует...")
    logger.info(f"[STARTUP] LOG_LEVEL        = {os.getenv('LOG_LEVEL', 'INFO')}")
    logger.info(f"[STARTUP] DATABASE_URL      = {os.getenv('DATABASE_URL','sqlite:////data/limits.db')}")
    logger.info(f"[STARTUP] ENTITY_TYPE_ID    = {os.getenv('BITRIX_ENTITY_TYPE_ID','139')}")
    logger.info(f"[STARTUP] CATEGORY_ID       = {os.getenv('BITRIX_CATEGORY_ID','38')}")
    logger.info(f"[STARTUP] FIELD_AMOUNT      = {os.getenv('FIELD_AMOUNT','UF_CRM_5_1698135335')}")
    logger.info(f"[STARTUP] FIELD_CURRENCY    = {os.getenv('FIELD_CURRENCY','UF_CRM_5_1698136041')}")
    logger.info(f"[STARTUP] FIELD_DEPARTMENT  = {os.getenv('FIELD_DEPARTMENT','(не задан!)')}")
    b24_url = os.getenv("BITRIX_WEBHOOK_URL", "")
    if b24_url and "YOUR" not in b24_url:
        logger.info(f"[STARTUP] BITRIX URL        = {b24_url[:40]}...")
    else:
        logger.warning("[STARTUP] BITRIX_WEBHOOK_URL не задан — интеграция с Б24 отключена")

    logger.info("[STARTUP] Инициализация БД...")
    init_db()

    logger.info("[STARTUP] Загрузка тестовых данных...")
    seed_test_data()

    logger.info("[STARTUP] Загрузка справочника валют из Б24...")
    try:
        mapping = await bitrix.load_currency_enum()
        if mapping:
            logger.info(f"[STARTUP] Валюты загружены: {mapping}")
        else:
            logger.warning("[STARTUP] Справочник валют пустой — проверьте BITRIX_WEBHOOK_URL")
    except Exception as e:
        logger.warning(f"[STARTUP] Справочник валют не загружен: {e}")

    logger.info("[STARTUP] ✓ Сервис готов к работе")
    logger.info("=" * 60)
    yield
    logger.info("[SHUTDOWN] Сервис остановлен")


app = FastAPI(
    title="Limit Service — Контроль лимитов Б24",
    version="1.2.0",
    lifespan=lifespan,
)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
templates = Jinja2Templates(directory="templates")


# ── Middleware: логируем каждый запрос ──────────────────────────
@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = datetime.utcnow()
    logger.debug(f"[REQ] {request.method} {request.url.path}")
    response = await call_next(request)
    elapsed = (datetime.utcnow() - start).total_seconds() * 1000
    level = logging.WARNING if response.status_code >= 400 else logging.DEBUG
    logger.log(level, f"[RES] {request.method} {request.url.path} → {response.status_code} ({elapsed:.0f}ms)")
    return response


# ────────────────────────────────────────────────────────────────
# PYDANTIC SCHEMAS
# ────────────────────────────────────────────────────────────────
class DepartmentCreate(BaseModel):
    name: str
    bitrix_department_value: Optional[str] = None
    description: Optional[str] = None

class DepartmentUpdate(BaseModel):
    name: Optional[str] = None
    bitrix_department_value: Optional[str] = None
    description: Optional[str] = None

class LimitUpsert(BaseModel):
    department_id: int
    year: int
    currency: str = "TJS"
    limit_amount: float

class ManualAdjust(BaseModel):
    department_id: int
    year: int
    currency: str
    amount: float
    note: Optional[str] = None


# ────────────────────────────────────────────────────────────────
# WEB UI
# ────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


# ────────────────────────────────────────────────────────────────
# ПОДРАЗДЕЛЕНИЯ
# ────────────────────────────────────────────────────────────────
@app.get("/api/departments")
def get_departments(year: int = Query(default=None), db: Session = Depends(get_db)):
    if year is None:
        year = datetime.now().year
    depts = db.query(Department).order_by(Department.name).all()
    logger.debug(f"[API] GET departments: {len(depts)} записей, год={year}")
    return [
        {
            "id": d.id, "name": d.name,
            "bitrix_department_value": d.bitrix_department_value,
            "description": d.description,
            "limits": [
                {
                    "id": l.id, "year": l.year, "currency": l.currency,
                    "limit_amount": l.limit_amount, "used_amount": l.used_amount,
                    "reserved_amount": l.reserved_amount,
                    "available_amount": l.available_amount,
                    "used_percent": l.used_percent,
                }
                for l in d.limits if l.year == year
            ]
        }
        for d in depts
    ]


@app.post("/api/departments", status_code=201)
def create_department(data: DepartmentCreate, db: Session = Depends(get_db)):
    logger.info(f"[API] POST department: name={data.name}, b24_val={data.bitrix_department_value}")
    if db.query(Department).filter_by(name=data.name).first():
        raise HTTPException(400, "Подразделение с таким именем уже существует")
    dept = Department(**data.model_dump())
    db.add(dept); db.commit(); db.refresh(dept)
    logger.info(f"[API] ✓ Создано подразделение id={dept.id}")
    return {"id": dept.id, "name": dept.name}


@app.put("/api/departments/{dept_id}")
def update_department(dept_id: int, data: DepartmentUpdate, db: Session = Depends(get_db)):
    dept = db.query(Department).filter_by(id=dept_id).first()
    if not dept:
        raise HTTPException(404, "Подразделение не найдено")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(dept, k, v)
    dept.updated_at = datetime.utcnow()
    db.commit()
    logger.info(f"[API] ✓ Обновлено подразделение id={dept_id}")
    return {"ok": True}


@app.delete("/api/departments/{dept_id}")
def delete_department(dept_id: int, db: Session = Depends(get_db)):
    dept = db.query(Department).filter_by(id=dept_id).first()
    if not dept:
        raise HTTPException(404, "Подразделение не найдено")
    logger.info(f"[API] DELETE department id={dept_id} name={dept.name}")
    db.delete(dept); db.commit()
    return {"ok": True}


# ────────────────────────────────────────────────────────────────
# ЛИМИТЫ
# ────────────────────────────────────────────────────────────────
@app.post("/api/limits")
def upsert_limit(data: LimitUpsert, db: Session = Depends(get_db)):
    dept = db.query(Department).filter_by(id=data.department_id).first()
    if not dept:
        raise HTTPException(404, "Подразделение не найдено")
    limit = db.query(DepartmentLimit).filter_by(
        department_id=data.department_id, year=data.year, currency=data.currency
    ).first()
    if limit:
        old = limit.limit_amount
        limit.limit_amount = data.limit_amount
        limit.updated_at = datetime.utcnow()
        note = f"Лимит изменён: {old:,.0f} → {data.limit_amount:,.0f} {data.currency}"
        logger.info(f"[LIMIT] {dept.name}: {note}")
    else:
        limit = DepartmentLimit(
            department_id=data.department_id, year=data.year,
            currency=data.currency, limit_amount=data.limit_amount,
        )
        db.add(limit)
        note = f"Лимит установлен: {data.limit_amount:,.0f} {data.currency}"
        logger.info(f"[LIMIT] {dept.name}: {note}")
    db.add(LimitTransaction(
        department_id=data.department_id, year=data.year, currency=data.currency,
        amount=data.limit_amount, transaction_type="manual_adjust",
        note=note, created_by="admin",
    ))
    db.commit()
    return {"ok": True, "message": note}


@app.post("/api/limits/adjust")
def manual_adjust(data: ManualAdjust, db: Session = Depends(get_db)):
    limit = db.query(DepartmentLimit).filter_by(
        department_id=data.department_id, year=data.year, currency=data.currency
    ).first()
    if not limit:
        raise HTTPException(404, "Лимит не найден")
    logger.info(f"[LIMIT] Ручная корректировка dept={data.department_id}: {data.amount:+,.0f} {data.currency}")
    limit.used_amount = max(0, limit.used_amount + data.amount)
    limit.updated_at = datetime.utcnow()
    db.add(LimitTransaction(
        department_id=data.department_id, year=data.year, currency=data.currency,
        amount=data.amount, transaction_type="manual_adjust",
        note=data.note or "Ручная корректировка", created_by="admin",
    ))
    db.commit()
    return {"ok": True}


# ────────────────────────────────────────────────────────────────
# ТРАНЗАКЦИИ
# ────────────────────────────────────────────────────────────────
@app.get("/api/transactions")
def get_transactions(
    department_id: Optional[int] = None,
    year: Optional[int] = None,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db)
):
    q = db.query(LimitTransaction)
    if department_id:
        q = q.filter_by(department_id=department_id)
    if year:
        q = q.filter_by(year=year)
    total = q.count()
    items = q.order_by(LimitTransaction.created_at.desc()).offset(offset).limit(limit).all()
    logger.debug(f"[API] GET transactions: total={total}, offset={offset}, limit={limit}")
    return {
        "total": total,
        "items": [
            {
                "id": t.id,
                "department": t.department.name if t.department else "—",
                "year": t.year, "currency": t.currency, "amount": t.amount,
                "type": t.transaction_type,
                "type_label": LimitTransaction.TYPE_LABELS.get(t.transaction_type, t.transaction_type),
                "bitrix_element_id": t.bitrix_element_id,
                "stage": t.bitrix_stage,
                "note": t.note,
                "created_at": t.created_at.strftime("%d.%m.%Y %H:%M") if t.created_at else "",
            }
            for t in items
        ]
    }


# ────────────────────────────────────────────────────────────────
# ЭКСПОРТ EXCEL
# ────────────────────────────────────────────────────────────────
@app.get("/api/export/excel")
def export_excel(year: int = Query(default=None), db: Session = Depends(get_db)):
    if year is None:
        year = datetime.now().year
    logger.info(f"[EXPORT] Excel выгрузка за {year} год")
    wb = openpyxl.Workbook()
    HF = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
    HFnt = openpyxl.styles.Font(bold=True, color="FFFFFF")

    ws1 = wb.active
    ws1.title = f"Лимиты {year}"
    ws1.append(["Подразделение", "Год", "Валюта", "Лимит", "Использовано", "Зарезервировано", "Остаток", "% исп."])
    for cell in ws1[1]:
        cell.fill = HF; cell.font = HFnt
    for d in db.query(Department).order_by(Department.name).all():
        for l in [x for x in d.limits if x.year == year]:
            ws1.append([d.name, l.year, l.currency, l.limit_amount,
                        l.used_amount, l.reserved_amount, l.available_amount, f"{l.used_percent}%"])
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 22

    ws2 = wb.create_sheet("Журнал транзакций")
    ws2.append(["ID", "Подразделение", "Год", "Валюта", "Сумма", "Тип", "Заявка Б24", "Стадия", "Примечание", "Дата"])
    for cell in ws2[1]:
        cell.fill = HF; cell.font = HFnt
    for t in db.query(LimitTransaction).filter_by(year=year).order_by(LimitTransaction.created_at.desc()).all():
        ws2.append([t.id, t.department.name if t.department else "—", t.year, t.currency, t.amount,
                    LimitTransaction.TYPE_LABELS.get(t.transaction_type, t.transaction_type),
                    t.bitrix_element_id or "—", t.bitrix_stage or "—", t.note or "—",
                    t.created_at.strftime("%d.%m.%Y %H:%M") if t.created_at else ""])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 22

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=limity_{year}.xlsx"})


# ────────────────────────────────────────────────────────────────
# НАСТРОЙКИ
# ────────────────────────────────────────────────────────────────
@app.get("/api/settings")
def get_settings(db: Session = Depends(get_db)):
    return {s.key: {"value": s.value, "description": s.description}
            for s in db.query(Setting).all()}

@app.post("/api/settings")
def update_settings(data: dict, db: Session = Depends(get_db)):
    for key, value in data.items():
        s = db.query(Setting).filter_by(key=key).first()
        if s:
            logger.info(f"[SETTINGS] {key} = '{value}'")
            s.value = value; s.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True}


# ────────────────────────────────────────────────────────────────
# БИТРИКС API
# ────────────────────────────────────────────────────────────────
@app.get("/api/bitrix/test")
async def test_bitrix():
    logger.info("[B24] Тест соединения...")
    result = await bitrix.test_connection()
    if result.get("ok"):
        logger.info(f"[B24] ✓ Соединение OK: {result.get('user', {}).get('NAME', '?')}")
    else:
        logger.warning(f"[B24] ✗ Соединение FAILED: {result.get('error')}")
    return result

@app.get("/api/bitrix/stages")
async def get_bitrix_stages():
    logger.info("[B24] Запрос стадий смарт-процесса...")
    stages = await bitrix.get_stages()
    logger.info(f"[B24] Получено стадий: {len(stages)}")
    for s in stages:
        logger.debug(f"[B24]   Стадия: ID={s.get('STATUS_ID')} NAME={s.get('NAME')}")
    return {"stages": stages, "category_id": bitrix.category_id}

@app.get("/api/bitrix/fields")
async def get_bitrix_fields():
    return {"fields": await bitrix.get_fields()}

@app.get("/api/bitrix/currencies")
async def get_bitrix_currencies():
    mapping = await bitrix.load_currency_enum()
    logger.info(f"[B24] Справочник валют обновлён: {mapping}")
    return {"currencies": mapping}

@app.get("/api/bitrix/categories")
async def get_bitrix_categories():
    return {"categories": await bitrix.get_categories()}


# ────────────────────────────────────────────────────────────────
# ВЕБХУК
# ────────────────────────────────────────────────────────────────
@app.post("/webhook/stage")
async def webhook_stage_change(request: Request, db: Session = Depends(get_db)):
    content_type = request.headers.get("content-type", "")
    element_id = None; stage_id = None
    raw_body = None

    try:
        if "application/json" in content_type:
            body = await request.json()
            raw_body = body
            element_id = body.get("element_id") or body.get("ID")
            stage_id   = body.get("stage_id") or body.get("STAGE_ID")
            secret     = body.get("secret", "")
            if secret and secret != WEBHOOK_SECRET:
                logger.warning(f"[WEBHOOK] ✗ Неверный секрет от {request.client.host}")
                raise HTTPException(403, "Неверный секрет")
        else:
            form = await request.form()
            body = dict(form)
            raw_body = body
            element_id = body.get("element_id") or body.get("data[FIELDS][ID]")
            stage_id   = body.get("stage_id") or body.get("data[FIELDS][STAGE_ID]")

            # Формат от робота Б24: document_id[2] = DYNAMIC_139_31410
            if not element_id:
                doc_id = body.get("document_id[2]", "")
                if doc_id:
                    parts = doc_id.split("_")
                    if parts:
                        element_id = parts[-1]
                        logger.info(f"[WEBHOOK] element_id из document_id[2]: {element_id}")
            secret     = body.get("secret", "")
            if secret and secret != WEBHOOK_SECRET:
                raise HTTPException(403, "Неверный секрет")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[WEBHOOK] ✗ Ошибка разбора тела запроса: {e}", exc_info=True)
        return JSONResponse({"ok": False, "error": "Ошибка разбора запроса"}, 400)

    if not element_id:
        logger.warning(f"[WEBHOOK] ✗ element_id не указан. Body: {raw_body}")
        return JSONResponse({"ok": False, "error": "element_id не указан"}, 400)

    element_id = int(element_id)
    logger.info(f"[WEBHOOK] ▶ Получен: element_id={element_id}, stage_id={stage_id!r}")

    settings_map = {s.key: s.value for s in db.query(Setting).all()}
    stage_reserve   = settings_map.get("stage_reserve", "")
    stage_execute   = settings_map.get("stage_execute", "")
    stage_release   = settings_map.get("stage_release", "")
    stage_unplanned = settings_map.get("stage_unplanned", "")
    field_dept      = settings_map.get("field_department", "")

    logger.debug(f"[WEBHOOK] Настройки: reserve={stage_reserve!r} execute={stage_execute!r} "
                 f"release={stage_release!r} unplanned={stage_unplanned!r}")

    if not stage_id:
        logger.info(f"[WEBHOOK] stage_id не передан, запрашиваем из Б24 для элемента {element_id}")
        element = await bitrix.get_element(element_id)
        if not element:
            logger.error(f"[WEBHOOK] ✗ Элемент {element_id} не найден в Б24")
            return JSONResponse({"ok": False, "error": "Элемент не найден"}, 404)
        stage_id = bitrix.extract_stage_id(element)
        logger.info(f"[WEBHOOK] Получена стадия из Б24: {stage_id!r}")
    else:
        element = None

    action = None
    if stage_id and stage_reserve  and stage_id == stage_reserve:  action = "reserve"
    elif stage_id and stage_execute and stage_id == stage_execute:  action = "execute"
    elif stage_id and stage_release and stage_id == stage_release:  action = "release"

    logger.info(f"[WEBHOOK] Определено действие: {action!r} (stage={stage_id!r})")

    if action is None:
        logger.info(f"[WEBHOOK] Стадия {stage_id!r} не требует обработки — пропускаем")
        return {"ok": True, "action": "skip", "stage_id": stage_id}

    if element is None:
        element = await bitrix.get_element(element_id)
        if not element:
            logger.error(f"[WEBHOOK] ✗ Не удалось загрузить элемент {element_id} из Б24")
            return JSONResponse({"ok": False, "error": "Элемент не найден"}, 404)

    amount       = bitrix.extract_amount(element)
    currency_raw = bitrix.extract_currency_raw(element)
    currency     = await bitrix.resolve_currency(currency_raw)
    dept_val     = bitrix.extract_department(element, field_dept)
    year         = datetime.now().year

    logger.info(f"[WEBHOOK] Данные элемента: amount={amount}, currency={currency}(raw={currency_raw}), "
                f"dept_val={dept_val!r}, year={year}")

    department = None
    if dept_val:
        department = db.query(Department).filter_by(bitrix_department_value=dept_val).first()
        if not department:
            department = db.query(Department).filter_by(name=dept_val).first()
        if department:
            logger.info(f"[WEBHOOK] Найдено подразделение: {department.name} (id={department.id})")
        else:
            logger.warning(f"[WEBHOOK] ✗ Подразделение '{dept_val}' не найдено в сервисе лимитов")

    # ── РЕЗЕРВИРОВАНИЕ ───────────────────────────────────────────
    if action == "reserve":
        if amount is None:
            logger.error(f"[WEBHOOK] ✗ Сумма заявки не определена для элемента {element_id}")
            return JSONResponse({"ok": False, "error": "Сумма заявки не определена"}, 400)

        if not department:
            _track(db, element_id, None, amount, currency, year, "error", stage_id,
                   f"Подразделение '{dept_val}' не найдено")
            await bitrix.add_comment(element_id,
                f"⚠️ Сервис лимитов: подразделение «{dept_val}» не найдено.\n"
                f"Обратитесь к администратору.")
            return {"ok": True, "action": "skip", "reason": "department_not_found"}

        limit = db.query(DepartmentLimit).filter_by(
            department_id=department.id, year=year, currency=currency
        ).first()

        if not limit:
            logger.warning(f"[WEBHOOK] ✗ Лимит {currency}/{year} не установлен для {department.name}")
            _track(db, element_id, department.id, amount, currency, year, "error", stage_id,
                   "Лимит не установлен")
            await bitrix.add_comment(element_id,
                f"⚠️ Лимит для «{department.name}» на {year} ({currency}) не установлен.")
            return {"ok": True, "action": "skip", "reason": "limit_not_set"}

        available = limit.available_amount
        logger.info(f"[WEBHOOK] Проверка лимита: нужно={amount:,.2f}, "
                    f"доступно={available:,.2f} (использ={limit.used_amount:,.2f}, "
                    f"резерв={limit.reserved_amount:,.2f}) {currency}")

        if available >= amount:
            limit.reserved_amount += amount
            limit.updated_at = datetime.utcnow()
            db.add(LimitTransaction(
                department_id=department.id, year=year, currency=currency,
                amount=amount, transaction_type="reserve",
                bitrix_element_id=element_id, bitrix_stage=stage_id,
                note=f"Резерв по заявке #{element_id}"
            ))
            _track(db, element_id, department.id, amount, currency, year, "reserved", stage_id)
            db.commit()
            logger.info(f"[WEBHOOK] ✓ РЕЗЕРВ: {amount:,.2f} {currency} для {department.name}. "
                        f"Остаток после: {available - amount:,.2f}")
            await bitrix.add_comment(element_id,
                f"✅ Лимит подтверждён\n"
                f"Подразделение: {department.name}\n"
                f"Сумма: {amount:,.2f} {currency}\n"
                f"Остаток лимита: {available - amount:,.2f} {currency}")
            return {"ok": True, "action": "reserved", "available_before": available, "amount": amount}

        else:
            logger.warning(f"[WEBHOOK] ✗ ЛИМИТ ПРЕВЫШЕН: нужно={amount:,.2f}, "
                           f"доступно={available:,.2f} {currency} [{department.name}]")
            _track(db, element_id, department.id, amount, currency, year, "unplanned", stage_id,
                   f"Лимит превышен: нужно {amount:,.2f}, доступно {available:,.2f}")
            db.add(LimitTransaction(
                department_id=department.id, year=year, currency=currency,
                amount=amount, transaction_type="unplanned",
                bitrix_element_id=element_id, bitrix_stage=stage_id,
                note=f"Лимит превышен: нужно {amount:,.2f}, доступно {available:,.2f} {currency}"
            ))
            db.commit()
            await bitrix.add_comment(element_id,
                f"⛔ ЛИМИТ ПРЕВЫШЕН\n"
                f"Подразделение: {department.name}\n"
                f"Запрошено: {amount:,.2f} {currency}\n"
                f"Доступно: {available:,.2f} {currency}\n"
                f"Заявка переведена на «Внеплановые расходы».\n"
                f"Требует согласования зам. по финансам.")
            if stage_unplanned:
                moved = await bitrix.update_stage(element_id, stage_unplanned)
                logger.info(f"[WEBHOOK] Перевод на внеплановые: {'OK' if moved else 'FAILED'}")
            return {"ok": True, "action": "moved_to_unplanned",
                    "needed": amount, "available": available}

    # ── ИСПОЛНЕНИЕ ───────────────────────────────────────────────
    elif action == "execute":
        tracking = db.query(ElementTracking).filter_by(bitrix_element_id=element_id).first()
        if tracking and tracking.status == "reserved" and tracking.department_id:
            limit = db.query(DepartmentLimit).filter_by(
                department_id=tracking.department_id,
                year=tracking.year or year, currency=tracking.currency
            ).first()
            if limit:
                limit.reserved_amount = max(0, limit.reserved_amount - tracking.amount)
                limit.used_amount += tracking.amount
                limit.updated_at = datetime.utcnow()
                db.add(LimitTransaction(
                    department_id=tracking.department_id, year=tracking.year or year,
                    currency=tracking.currency, amount=tracking.amount,
                    transaction_type="execute", bitrix_element_id=element_id,
                    bitrix_stage=stage_id, note=f"Исполнено: заявка #{element_id}"
                ))
            tracking.status = "executed"; tracking.last_stage = stage_id
            tracking.updated_at = datetime.utcnow()
            db.commit()
            logger.info(f"[WEBHOOK] ✓ ИСПОЛНЕНО: {tracking.amount:,.2f} {tracking.currency} "
                        f"для элемента #{element_id}")
            return {"ok": True, "action": "executed", "amount": tracking.amount}
        logger.warning(f"[WEBHOOK] Исполнение: нет резерва для элемента #{element_id}")
        return {"ok": True, "action": "skip", "reason": "no_reservation_found"}

    # ── ОТКЛОНЕНИЕ ───────────────────────────────────────────────
    elif action == "release":
        tracking = db.query(ElementTracking).filter_by(bitrix_element_id=element_id).first()
        if tracking and tracking.status == "reserved" and tracking.department_id:
            limit = db.query(DepartmentLimit).filter_by(
                department_id=tracking.department_id,
                year=tracking.year or year, currency=tracking.currency
            ).first()
            if limit:
                limit.reserved_amount = max(0, limit.reserved_amount - tracking.amount)
                limit.updated_at = datetime.utcnow()
                db.add(LimitTransaction(
                    department_id=tracking.department_id, year=tracking.year or year,
                    currency=tracking.currency, amount=tracking.amount,
                    transaction_type="release", bitrix_element_id=element_id,
                    bitrix_stage=stage_id, note=f"Резерв снят: заявка #{element_id} отклонена"
                ))
            tracking.status = "released"; tracking.last_stage = stage_id
            tracking.updated_at = datetime.utcnow()
            db.commit()
            logger.info(f"[WEBHOOK] ✓ РЕЗЕРВ СНЯТ: {tracking.amount:,.2f} для элемента #{element_id}")
            return {"ok": True, "action": "released", "amount": tracking.amount}
        logger.warning(f"[WEBHOOK] Снятие резерва: нет резерва для элемента #{element_id}")
        return {"ok": True, "action": "skip", "reason": "no_reservation_found"}


def _track(db, element_id, dept_id, amount, currency, year, status, stage, error=None):
    existing = db.query(ElementTracking).filter_by(bitrix_element_id=element_id).first()
    if existing:
        existing.status = status; existing.last_stage = stage
        existing.updated_at = datetime.utcnow()
        if error: existing.error_message = error
    else:
        db.add(ElementTracking(
            bitrix_element_id=element_id, department_id=dept_id,
            amount=amount, currency=currency, year=year,
            status=status, last_stage=stage, error_message=error,
        ))


# ────────────────────────────────────────────────────────────────
# ЗДОРОВЬЕ / СТАТИСТИКА
# ────────────────────────────────────────────────────────────────
@app.get("/health")
def health(db: Session = Depends(get_db)):
    depts = db.query(Department).count()
    txs   = db.query(LimitTransaction).count()
    return {
        "status": "ok", "service": "limit-service", "version": "1.2.0",
        "timestamp": datetime.now().isoformat(),
        "db": {"departments": depts, "transactions": txs},
        "config": {
            "entity_type_id": os.getenv("BITRIX_ENTITY_TYPE_ID", "139"),
            "category_id": os.getenv("BITRIX_CATEGORY_ID", "38"),
            "field_department_set": bool(os.getenv("FIELD_DEPARTMENT")),
            "bitrix_url_set": bool(os.getenv("BITRIX_WEBHOOK_URL")),
        }
    }


@app.get("/api/stats")
def get_stats(year: int = Query(default=None), db: Session = Depends(get_db)):
    if year is None:
        year = datetime.now().year
    limits = db.query(DepartmentLimit).filter_by(year=year).all()
    total_limit    = sum(l.limit_amount    for l in limits)
    total_used     = sum(l.used_amount     for l in limits)
    total_reserved = sum(l.reserved_amount for l in limits)
    total_available= sum(l.available_amount for l in limits)
    over_limit     = [l for l in limits if l.used_percent >= 90]
    return {
        "year": year,
        "total_limit": total_limit, "total_used": total_used,
        "total_reserved": total_reserved, "total_available": total_available,
        "departments_count": db.query(Department).count(),
        "transactions_count": db.query(LimitTransaction).filter_by(year=year).count(),
        "over_limit_count": len(over_limit),
        "used_percent": round((total_used + total_reserved) / total_limit * 100, 1) if total_limit else 0,
    }
